import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('test.xlsx')
sheet = wb['Sheet1']
print(sheet.max_row)
print(sheet['a1'].value)

for row in range(2, sheet.max_row + 1):
	corrected_value = sheet.cell(row, 3).value * 0.9
	corrected_cell = sheet.cell(row, 4)
	corrected_cell.value = corrected_value
wb.save('corrected_value.xlsx') 

values = Reference(sheet,
	min_row = 2,
	max_row = sheet.max_row,
	min_col = 3,
	max_col = 4)

chart = BarChart()
chart.add_data(values)

sheet.add_chart(chart, 'e2')
wb.save('corrected_value.xlsx') 
